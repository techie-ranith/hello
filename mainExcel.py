import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os

url = "https://autostream.lk/listings/honda-vezel-2024-49/"
response = requests.get(url)
soup = BeautifulSoup(response.text, "html.parser")

data = {}

# 1️⃣ Main attributes (Body, Mileage, Fuel Type, Engine CC, etc.)
for item in soup.select(".single-listing-attribute-boxes .item"):
    label = item.select_one(".label-text")
    value = item.select_one(".value-text")
    if label and value:
        data[label.get_text(strip=True)] = value.get_text(strip=True)

print("📊 Main Attributes:")
for key, value in data.items():
    print(f"  {key}: {value}")
print()

# 2️⃣ Data list (Grade, District, Exterior/Interior Color, etc.)
for li in soup.select(".stm-single-car-listing-data .data-list-item"):
    label = li.select_one(".item-label")
    value = li.select_one(".heading-font")
    if label and value:
        data[label.get_text(strip=True)] = value.get_text(strip=True)

print("📋 Additional Data:")
for key, value in list(data.items())[-5:]:  # Show last 5 items added
    print(f"  {key}: {value}")
print()

# 3️⃣ Features (grouped by category)
features = {}
for group in soup.select(".stm-single-listing-car-features .grouped_checkbox-3"):
    category = group.select_one("h4")
    if not category:
        continue
    category_name = category.get_text(strip=True)
    features_list = [li.get_text(strip=True) for li in group.select("ul li span") if li.get_text(strip=True)]
    features[category_name] = features_list

print("🚗 Features by Category:")
for category, feature_list in features.items():
    print(f"  {category}: {', '.join(feature_list)}")
print()

# Flatten features into a single string for Excel
for cat, feats in features.items():
    data[cat] = ", ".join(feats)

# 4️⃣ Seller Notes
seller_notes = soup.select_one("section:has(h2:contains('Seller Notes'))")
if seller_notes:
    data["Seller Notes"] = seller_notes.get_text(strip=True)
    print("📝 Seller Notes:")
    print(f"  {data['Seller Notes']}")
    print()

print("=" * 50)
print("📋 COMPLETE DATA SUMMARY:")
print("=" * 50)
for key, value in data.items():
    print(f"{key}: {value}")
print("=" * 50)

# ✅ Save / Append to Excel
file_name = "vehicle_data.xlsx"

if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    # Write header
    ws.append(list(data.keys()))

# Append row
ws.append(list(data.values()))

# Save workbook
wb.save(file_name)
print(f"✅ Data appended to {file_name}")
