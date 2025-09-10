import time
import os
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

BASE_URL = "https://autostream.lk"


# --------------------------
# Normalize scraped labels to Excel header keys
# --------------------------
def normalize_field_name(label: str) -> str | None:
    key = label.lower()
    key = re.sub(r'[\s\-\:_()]+', ' ', key).strip()
    key = key.replace('k w', 'kw').replace('cc / kw', 'cc/kw').replace('cc /kw', 'cc/kw').replace('cc/ kw', 'cc/kw')
    mapping = {
        'fuel type': 'Fuel Type',
        'fuel': 'Fuel Type',
        'engine cc/kw': 'Engine CC / kw',
        'engine cc kw': 'Engine CC / kw',
        'engine cckw': 'Engine CC / kw',
        'engine cc': 'Engine CC / kw',
        'engine capacity': 'Engine CC / kw',
        'engine capacity cc': 'Engine CC / kw',
        'engine': 'Engine CC / kw',
    }
    return mapping.get(key)


# --------------------------
# Extract dealer info (from dealer page)
# --------------------------
def extract_dealer_info_from_dealer_page(soup: BeautifulSoup) -> dict:
    info = {
        "Dealer Name": "",
        "Dealership Location": "",
        "Sales Hours": "",
        "Seller Email": "",
        "Dealer Contact Number": "",
    }
    # Dealer name
    el = soup.select_one("h1, h2, h3, h4, .dealer-title, .dealer-name, .name, .title, .author-title")
    if el: info["Dealer Name"] = el.get_text(strip=True)

    # Location
    el = soup.select_one(".stm-dealer-location, .dealer-location, .location, .dealer-address, address")
    if el: info["Dealership Location"] = el.get_text(strip=True)

    # Sales Hours
    el = soup.select_one(".dealer-working-hours, .working-hours, .hours")
    if el: info["Sales Hours"] = el.get_text(strip=True)

    # Email
    mail_el = soup.select_one("a[href^='mailto:']")
    if mail_el: info["Seller Email"] = mail_el.get_text(strip=True) or mail_el["href"].split(":")[1]

    # Phone
    phone_el = soup.select_one("a[href^='tel:']")
    if phone_el: info["Dealer Contact Number"] = phone_el.get_text(strip=True) or phone_el["href"].split(":")[1]

    return info


# --------------------------
# Scrape single vehicle ad
# --------------------------
def scrape_vehicle(url, dealer_info):
    print(f"🔹 Scraping vehicle: {url}")
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    data = {**dealer_info, "Ad URL": url}

    # Vehicle name + price
    vehicle_name = soup.select_one("h1.listing-title, h6.title.stm_listing_title")
    data["Vehicle Name"] = vehicle_name.get_text(strip=True) if vehicle_name else ""

    vehicle_price = soup.select_one(".price .heading-font, span.h3")
    data["Vehicle Price"] = vehicle_price.get_text(strip=True) if vehicle_price else ""

    # Vehicle status
    status_tag = soup.select_one("div.special-label.h5")
    if status_tag and "sold" in status_tag.get_text(strip=True).lower():
        data["Status"] = "Sold"
    else:
        data["Status"] = "Available"

    # Attributes
    for item in soup.select(".single-listing-attribute-boxes .item"):
        label = item.select_one(".label-text")
        value = item.select_one(".value-text")
        if label:
            label_text = label.get_text(strip=True)
            value_text = value.get_text(strip=True) if value else ""
            canonical = normalize_field_name(label_text)
            key = canonical if canonical else label_text
            data[key] = value_text

    for li in soup.select(".stm-single-car-listing-data .data-list-item"):
        label = li.select_one(".item-label")
        value = li.select_one(".heading-font")
        if label and value:
            label_text = label.get_text(strip=True)
            canonical = normalize_field_name(label_text)
            key = canonical if canonical else label_text
            data[key] = value.get_text(strip=True)

    # Features
    for group in soup.select(".stm-single-listing-car-features .grouped_checkbox-3"):
        category = group.select_one("h4")
        if category:
            features_list = [li.get_text(strip=True) for li in group.select("ul li span")]
            data[category.get_text(strip=True)] = ", ".join(features_list)

    return data


# --------------------------
# Scrape all ads from one dealer
# --------------------------
def scrape_dealer(dealer_url):
    ads = []

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get(dealer_url)

    # click "Show more" until all ads loaded
    while True:
        try:
            show_more = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@class='heading-font']/span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'show more')]/.."))
            )
            print("🔘 Clicking Show more...")
            driver.execute_script("arguments[0].click();", show_more)
            time.sleep(2)
        except Exception:
            break

    soup = BeautifulSoup(driver.page_source, "html.parser")
    dealer_info = extract_dealer_info_from_dealer_page(soup)

    # Collect ad links
    ad_links = []
    for listing in soup.select(".car-listing-row.row.row-3"):
        for a_tag in listing.find_all("a", href=True):
            if "/listings/" in a_tag["href"]:
                ad_links.append(a_tag["href"])
    driver.quit()

    # Remove duplicates
    ad_links = list(dict.fromkeys(ad_links))
    print(f"🔎 Found {len(ad_links)} ads for dealer {dealer_info.get('Dealer Name', '')}")

    # Scrape each ad
    for ad_url in ad_links:
        if not ad_url.startswith("http"):
            ad_url = BASE_URL + ad_url
        try:
            ads.append(scrape_vehicle(ad_url, dealer_info))
            time.sleep(1)
        except Exception as e:
            print(f"❌ Failed to scrape {ad_url}: {e}")

    return ads


# --------------------------
# Get all dealers
# --------------------------
# --------------------------
# Get all dealers (with "Show more")
# --------------------------
def get_dealers():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get(f"{BASE_URL}/dealers/")

    # Keep clicking "Show more" until none left
    while True:
        try:
            show_more = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@class='stm-load-more-dealers button']/span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'show more')]/.."))
            )
            print("🔘 Clicking Show more (dealers)...")
            driver.execute_script("arguments[0].click();", show_more)
            time.sleep(2)
        except Exception:
            print("✅ All dealers loaded.")
            break

    # Parse full dealer list
    soup = BeautifulSoup(driver.page_source, "html.parser")
    dealers = []
    for row in soup.select("tr.stm-single-dealer"):
        a = row.select_one(".dealer-info .h4")
        if a and a.get("href"):
            dealer_name = a.get_text(strip=True)
            dealer_url = a["href"]
            if not dealer_url.startswith("http"):
                dealer_url = BASE_URL + dealer_url
            print(f"📌 Dealer found: {dealer_name} -> {dealer_url}")
            dealers.append(dealer_url)

    driver.quit()
    return dealers


def sort_by_alphabet(driver):
    # Click the dropdown
    dropdown = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "select2-oxzl-container"))
    )
    dropdown.click()

    # Wait for options to appear and click "Alphabet"
    alphabet_option = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//li[contains(text(), 'Alphabet')]"))
    )
    alphabet_option.click()
    print("✅ Sorted dealers by Alphabet")


# --------------------------
# Save to Excel
# --------------------------
def save_to_excel(data_list, file_name="vehicle_data.xlsx"):
    if not data_list:
        return
    headers = [
        "Dealer Name", "Dealership Location", "Sales Hours", "Seller Email", "Dealer Contact Number",
        "Vehicle Name", "Vehicle Price", "Status", "Body", "Mileage", "Fuel Type", "Engine CC / kw",
        "Year of Manufacture", "Transmission", "Grade", "Exterior Color", "Interior Color",
        "No. of Owners", "District", "City", "Year of Reg.", "Convenience", "Infotainment",
        "Safety & Security", "Interior & Seats", "Windows & Lighting", "Other Features",
        "Seller Notes", "Ad URL"
    ]

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    for row in data_list:
        ws.append([row.get(h, "") for h in headers])

    wb.save(file_name)


# --------------------------
# MAIN PROCESS
# --------------------------
if __name__ == "__main__":
    all_ads = []
    dealers = get_dealers()
    print(f"🌐 Found {len(dealers)} dealers")
    for dealer_url in dealers:
        try:
            dealer_ads = scrape_dealer(dealer_url)
            all_ads.extend(dealer_ads)
        except Exception as e:
            print(f"❌ Failed to scrape dealer {dealer_url}: {e}")

    save_to_excel(all_ads)
    print(f"✅ Scraping complete! {len(all_ads)} ads saved to Excel.")
