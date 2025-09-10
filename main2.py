import time
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re

BASE_URL = "https://autostream.lk"


# --------------------------
# Normalize scraped labels to Excel header keys
# --------------------------
def normalize_field_name(label: str) -> str | None:
    # unify spacing, punctuation, and case
    key = label.lower()
    key = re.sub(r'[\s\-\:_()]+', ' ', key).strip()
    key = key.replace('k w', 'kw')
    key = key.replace('cc / kw', 'cc/kw')
    key = key.replace('cc /kw', 'cc/kw')
    key = key.replace('cc/ kw', 'cc/kw')

    mapping = {
        # Fuel
        'fuel type': 'Fuel Type',
        'fuel': 'Fuel Type',
        # Engine CC / kW variants
        'engine cc/kw': 'Engine CC / kw',
        'engine cc kw': 'Engine CC / kw',
        'engine cckw': 'Engine CC / kw',
        'engine cc': 'Engine CC / kw',
        'engine capacity': 'Engine CC / kw',
        'engine capacity cc': 'Engine CC / kw',
        'engine': 'Engine CC / kw',
    }
    return mapping.get(key)

# Fill missing dealer fields from ad page
def _extract_text_or_href(el, scheme_prefix: str) -> str:
    if not el:
        return ""
    txt = el.get_text(strip=True)
    if txt:
        return txt
    href = el.get("href", "")
    if href and href.startswith(scheme_prefix):
        return href.split(":", 1)[1]
    return ""

def extract_dealer_info_from_dealer_page(soup: BeautifulSoup) -> dict:
    # Try common blocks on dealer/author pages
    block = soup.select_one(
        ".stm-dealer-info, .dealer-info, .stm-dealer-box, .stm-dealer-details, "
        ".author-info, .stm-seller-info, .seller-info, .dealer-contact"
    )
    info = {
        "Dealer Name": "",
        "Dealership Location": "",
        "Sales Hours": "",
        "Seller Email": "",
        "Dealer Contact Number": "",
    }

    # Dealer name
    if block:
        el = block.select_one("h1, h2, h3, h4, .dealer-title, .dealer-name, .name, .title")
        if el:
            info["Dealer Name"] = el.get_text(strip=True)

    if not info["Dealer Name"]:
        el = soup.select_one("h1, .page-title, .entry-title, .author-title")
        if el:
            info["Dealer Name"] = el.get_text(strip=True)

    # Location
    if block and not info["Dealership Location"]:
        el = block.select_one(".stm-dealer-location, .dealer-location, .location, .dealer-address, address")
        if el:
            info["Dealership Location"] = el.get_text(strip=True)
    if not info["Dealership Location"]:
        # heuristic search by label
        label = soup.find(string=re.compile(r"location", re.I))
        if label and label.parent:
            info["Dealership Location"] = label.parent.get_text(" ", strip=True).replace("Location", "").strip()

    # Sales Hours
    if block:
        el = block.select_one(".dealer-working-hours, .working-hours, .hours, .dealer-hours")
        if el:
            info["Sales Hours"] = el.get_text(strip=True)
    if not info["Sales Hours"]:
        label = soup.find(string=re.compile(r"(sales|working)\s*hours", re.I))
        if label and label.parent:
            info["Sales Hours"] = label.parent.get_text(" ", strip=True)

    # Email/Phone (global)
    mail_el = soup.select_one("a[href^='mailto:']")
    info["Seller Email"] = _extract_text_or_href(mail_el, "mailto:")
    phone_el = soup.select_one("a[href^='tel:']")
    info["Dealer Contact Number"] = _extract_text_or_href(phone_el, "tel:")
    if not info["Dealer Contact Number"]:
        txt = soup.get_text(" ", strip=True)
        m = re.search(r"\+?\d[\d\s\-\(\)]{7,}", txt)
        if m:
            info["Dealer Contact Number"] = m.group(0).strip()

    return info

def enrich_dealer_from_ad_page(soup, data: dict):
    dealer_block = soup.select_one(".dealer-info, .stm-dealer-box, .stm-dealer-info, .stm-seller-info")
    if not dealer_block:
        return
    if not data.get("Dealer Name"):
        el = dealer_block.select_one("h3, .dealer-title, .name, h4")
        data["Dealer Name"] = el.get_text(strip=True) if el else ""
    if not data.get("Dealership Location"):
        el = dealer_block.select_one(".stm-dealer-location, .dealer-location, .location")
        data["Dealership Location"] = el.get_text(strip=True) if el else ""
    if not data.get("Sales Hours"):
        el = dealer_block.select_one(".dealer-working-hours, .working-hours")
        data["Sales Hours"] = el.get_text(strip=True) if el else ""
    if not data.get("Seller Email"):
        el = dealer_block.select_one("a[href^='mailto:']")
        data["Seller Email"] = _extract_text_or_href(el, "mailto:")
    if not data.get("Dealer Contact Number"):
        el = dealer_block.select_one("a[href^='tel:']")
        data["Dealer Contact Number"] = _extract_text_or_href(el, "tel:")

# --------------------------
# Scraper function for a single vehicle ad
# --------------------------
def scrape_vehicle(url, dealer_info):
    print(f"🔹 Scraping vehicle: {url}")
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    # start with dealer info (copied into every vehicle row)
    data = {**dealer_info, "Ad URL": url}

    # Vehicle details
    vehicle_name = soup.select_one("h1.listing-title, h6.title.stm_listing_title")
    data["Vehicle Name"] = vehicle_name.get_text(strip=True) if vehicle_name else ""

    vehicle_price = soup.select_one(".price .heading-font, span.h3")
    data["Vehicle Price"] = vehicle_price.get_text(strip=True) if vehicle_price else ""
    
    # Vehicle status (Sold / Available)
    status_tag = soup.select_one("div.special-label.h5")
    if status_tag and "sold" in status_tag.get_text(strip=True).lower():
        data["Status"] = "Sold"
    else:
        data["Status"] = "Available"


    # Main attributes (Body, Mileage, Fuel Type, Engine CC)
    for item in soup.select(".single-listing-attribute-boxes .item"):
        label = item.select_one(".label-text")
        value = item.select_one(".value-text")
        if label:
            label_text = label.get_text(strip=True)
            value_text = value.get_text(strip=True) if value else item.get_text(strip=True).replace(label_text, "").strip()
            canonical = normalize_field_name(label_text)
            key = canonical if canonical else label_text
            data[key] = value_text

    # Additional attributes
    for li in soup.select(".stm-single-car-listing-data .data-list-item"):
        label = li.select_one(".item-label")
        value = li.select_one(".heading-font")
        if label and value:
            label_text = label.get_text(strip=True)
            canonical = normalize_field_name(label_text)
            key = canonical if canonical else label_text
            data[key] = value.get_text(strip=True)

    # Features
    for group in soup.select(".stm-single-listing-car-features .grouped_checkbox-3"):
        category = group.select_one("h4")
        if not category:
            continue
        category_name = category.get_text(strip=True)
        features_list = [li.get_text(strip=True) for li in group.select("ul li span") if li.get_text(strip=True)]
        data[category_name] = ", ".join(features_list)

    # Seller Notes
    seller_notes = soup.select_one("section:has(h2:-soup-contains('Seller Notes'))")
    if seller_notes:
        data["Seller Notes"] = seller_notes.get_text(strip=True)

    # Ensure dealer fields are filled
    enrich_dealer_from_ad_page(soup, data)

    return data

# --------------------------
# Collect all ads from a dealer (scrapes dealer info once)
# --------------------------
def scrape_dealer(dealer_url):
    ads = []

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get(dealer_url)

    while True:
        try:
            show_more = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((
                    By.XPATH, 
                    "//a[@class='heading-font']/span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'show more')]/.."
                ))
            )
            print("🔘 Clicking Show more...")
            driver.execute_script("arguments[0].click();", show_more)
            time.sleep(2)
        except Exception:
            print("✅ No more Show more button.")
            break

    # Page source after expanding
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 1) Extract and store dealer details first (dealer-only row)
    dealer_info = extract_dealer_info_from_dealer_page(soup)
    dealer_row = {
        "Dealer Name": dealer_info.get("Dealer Name", ""),
        "Dealership Location": dealer_info.get("Dealership Location", ""),
        "Sales Hours": dealer_info.get("Sales Hours", ""),
        "Seller Email": dealer_info.get("Seller Email", ""),
        "Dealer Contact Number": dealer_info.get("Dealer Contact Number", ""),
        "Ad URL": dealer_url,  # reference to dealer page
    }
    ads.append(dealer_row)

    # 2) Collect ad links
    ad_links = []
    for listing in soup.select(".car-listing-row.row.row-3"):
        for a_tag in listing.find_all("a", href=True):
            if "/listings/" in a_tag["href"]:
                ad_links.append(a_tag["href"])

    driver.quit()

    ad_links = list(set(ad_links))
    print(f"🔎 Found {len(ad_links)} ads in total")

    # 3) Scrape each ad with dealer_info injected
# 3) Scrape each ad with dealer_info injected
    for ad_url in ordered_links:
        if not ad_url.startswith("http"):
            ad_url = BASE_URL + ad_url
        try:
            ad_data = scrape_vehicle(ad_url, dealer_info)
            ads.append(ad_data)
            time.sleep(1)
        except Exception as e:
            print(f"❌ Failed to scrape {ad_url}: {e}")


    print(f"✅ Total ads scraped: {len(ads)}")
    return ads


# --------------------------
# Save data to Excel
# --------------------------
def save_to_excel(data_list, file_name="vehicle_data3.xlsx"):
    if not data_list:
        return

    headers = [
        "Dealer Name", "Dealership Location", "Sales Hours", "Seller Email", "Dealer Contact Number",
        "Vehicle Name", "Vehicle Price", "Status", "Contact Number", "Registration Number",
        "Body", "Mileage", "Fuel Type", "Engine CC / kw", "Year of Manufacture", "Transmission",
        "Grade", "Exterior Color", "Interior Color", "No. of Owners", "Blue-T Grade",
        "District", "City", "Year of Reg.", "Convenience", "Infotainment", "Safety & Security",
        "Interior & Seats", "Windows & Lighting", "Other Features", "Seller Notes", "Ad URL"
    ]


    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        # ensure header row matches
        existing = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if existing != headers:
            for i, h in enumerate(headers, start=1):
                ws.cell(row=1, column=i, value=h)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    for row in data_list:
        ws.append([row.get(h, "") for h in headers])

    wb.save(file_name)

# --------------------------
# MAIN
# --------------------------
dealer_url = "https://autostream.lk/author/dvithanageyahoo-com/"
ads_data = scrape_dealer(dealer_url)
save_to_excel(ads_data)